"""
阅读excel当中内容，并且提取指定内容到txt当中
"""

import numpy as np
from openpyxl import load_workbook


def read_excel(path, rows=0, cols=0, data_only=True, ):
    """
    读取excel表格
    :param path:xlsx文件路径
    :param rows: 读取的表格行范围,如果为负数,则从后记
    :param cols: 读取的表格列范围,如果为负数,则从后记
    :param data_only: 仅读取其中数据
    :return: excel表格当中内容
    """
    Workbook = load_workbook(path, data_only=data_only)

    Worksheet = Workbook[Workbook.sheetnames[1]]
    if rows > 0:
        rows = rows
    elif rows <= 0:
        rows = Worksheet.max_row + rows

    if cols > 0:
        cols = cols
    elif cols <= 0:
        cols = Worksheet.max_column + cols
    data_name = [name[0] for name in Worksheet.iter_cols(min_col=1, max_col=cols, min_row=1, max_row=1, values_only=True)]
    # print(data_name)
    data = [dict(zip(data_name, list(row))) for row in Worksheet.iter_rows(min_col=1, max_col=cols, min_row=2, max_row=rows, values_only=True)]
    # for row in Worksheet.iter_rows(min_col=1, max_col=cols, min_row=2, max_row=rows, values_only=True):
    #     print(list(row))
    # print(np.array(data))
    return np.array(data_name), np.array(data)


def peoples_handle(peoples):
    peoples = peoples.replace(u'\xa0', u' ')
    return ';'.join([people[2:] for people in peoples.split(' ')])


def case_1(path, rows, cols, ):
    """
    需要更具需求来进行重写
    :param path:xlsx文件路径
    :param rows:读取的excel表格范围
    :param cols:读取的excel表格范围
    :return:保存在当前路径下,格式为txt
    """
    txt_path = path.replace('xlsx', 'txt')
    data_name, data_list = read_excel(path, rows=rows, cols=cols)
    print(data_name)
    with open(txt_path, 'w', encoding='utf-8') as f:
        for i, data in enumerate(data_list):
            index = i + 1
            peoples = str(peoples_handle(data['发明人']))
            name = data['专利名称']
            try:
                time = '{:%Y.%m.%d}'.format(data['授权日期'])
            except:
                time = data['授权日期']
            nation = '中国'
            number = data['专利号']
            # print(f'{index}.\t{peoples},{name},{time},{nation},{number}')
            f.write(f'{index}.\t{peoples},{name},{time},{nation},{number}\n\n')


if __name__ == '__main__':
    path = r"D:\Files\GitHub\Utils\temp\王老师专利申请表及授权汇总(2).xlsx"
    txt_path = path.replace('xlsx', 'txt')
    rows = -12
    cols = -4
    data_name, data_list = read_excel(path, rows=rows, cols=cols)
    print(data_name)
    with open(txt_path, 'w', encoding='utf-8') as f:
        for i, data in enumerate(data_list):
            index = i + 1
            peoples = str(peoples_handle(data['发明人']))
            name = data['专利名称']
            try:
                time = '{:%Y.%m.%d}'.format(data['授权日期'])
            except:
                time = data['授权日期']
            nation = '中国'
            number = data['专利号']
            # print(f'{index}.\t{peoples},{name},{time},{nation},{number}')
            f.write(f'{index}.\t{peoples},{name},{time},{nation},{number}\n\n')
